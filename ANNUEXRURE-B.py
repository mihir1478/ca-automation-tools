import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
import io

# --- Page Settings ---
st.set_page_config(page_title="GSTR-2B Auto Filler", page_icon="🖨️", layout="centered")

st.title("🖨️ GSTR-2B to Annexure-B (Calibri & Smart Trim)")
st.write("Ab 'PARTNER' ka naam 100% safe rahega aur uske theek 2 line baad sheet cut hogi!")
st.markdown("---")

col1, col2 = st.columns(2)

with col1:
    st.subheader("📁 Step 1")
    template_file = st.file_uploader("Upload Annexure-B (.xlsx file)", type=['xlsx'])

with col2:
    st.subheader("📁 Step 2")
    gstr2b_file = st.file_uploader("Upload GSTR-2B File (Govt. Portal)", type=['xlsx', 'xls'])

st.markdown("---")

if st.button("🚀 Process Data & Generate File", type="primary", use_container_width=True):
    
    if template_file is None or gstr2b_file is None:
        st.warning("⚠️ Kripya process karne se pehle dono files upload karein!")
    else:
        with st.spinner("⏳ Data process ho raha hai aur Partner ka naam save kiya ja raha hai..."):
            try:
                # --- 1. GSTR-2B Smart Header Finder ---
                df_raw = pd.read_excel(gstr2b_file, sheet_name='B2B', header=None)
                header_idx = 5 
                for i, r in df_raw.head(15).iterrows():
                    vals = str(r.values).lower()
                    if 'gstin' in vals and ('trade' in vals or 'legal' in vals or 'name' in vals):
                        header_idx = i
                        break
                        
                df_gstr2b = pd.read_excel(gstr2b_file, sheet_name='B2B', header=header_idx)
                df_gstr2b = df_gstr2b.fillna('')
                
                cols = [str(c).lower().replace(' ', '').replace('\n', '').replace('/', '') for c in df_gstr2b.columns]
                df_gstr2b.columns = cols

                # --- 2. Accurate Column Extraction ---
                col_gstin = next((c for c in cols if 'gstin' in c), cols[0] if len(cols)>0 else None)
                col_name = next((c for c in cols if 'trade' in c or 'legal' in c or 'suppliername' in c), cols[1] if len(cols)>1 else None)
                col_inv = next((c for c in cols if 'invoiceno' in c or 'invoicenumber' in c), cols[2] if len(cols)>2 else None)
                col_date = next((c for c in cols if 'date' in c), cols[4] if len(cols)>4 else None)
                
                col_taxable = next((c for c in cols if 'taxable' in c), cols[9] if len(cols)>9 else None)
                col_igst = next((c for c in cols if 'integrated' in c or 'igst' in c), cols[10] if len(cols)>10 else None)
                col_cgst = next((c for c in cols if 'central' in c or 'cgst' in c), cols[11] if len(cols)>11 else None)
                col_sgst = next((c for c in cols if 'state' in c or 'sgst' in c), cols[12] if len(cols)>12 else None)
                col_cess = next((c for c in cols if 'cess' in c), cols[13] if len(cols)>13 else None)

                def get_amount(val):
                    try:
                        return float(str(val).replace(',', '').strip()) if str(val).strip() != '' else 0.0
                    except: return 0.0

                # --- 3. Annexure-B File Open & Setup ---
                wb = openpyxl.load_workbook(template_file)
                sheet = wb.active

                # 🖨️ PRINT SETUP
                sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
                sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
                sheet.sheet_properties.pageSetUpPr.fitToPage = True
                sheet.page_setup.fitToWidth = 1
                sheet.page_setup.fitToHeight = 0 
                sheet.page_margins = PageMargins(left=0.2, right=0.2, top=0.5, bottom=0.5, header=0.2, footer=0.2)
                sheet.print_title_rows = '1:5'

                # 📏 COLUMN WIDTHS FIX
                sheet.column_dimensions['A'].width = 5 
                sheet.column_dimensions['B'].width = 17  
                sheet.column_dimensions['C'].width = 30  
                sheet.column_dimensions['D'].width = 16  
                sheet.column_dimensions['E'].width = 12  
                for col in ['J', 'M', 'N', 'O', 'P', 'R']:    
                    sheet.column_dimensions[col].width = 14

                # 🔥 STYLES SETUP (Calibri Font)
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
                align_left = Alignment(horizontal='left', vertical='center', wrap_text=True) 
                align_right = Alignment(horizontal='right', vertical='center')
                
                normal_font = Font(name='Calibri', size=11, bold=False) 
                bold_font = Font(name='Calibri', size=11, bold=True)

                start_row = 6 
                num_rows = len(df_gstr2b)

                sheet.insert_rows(start_row, amount=num_rows)

                # --- 4. Data Transfer Loop & TOTAL CALCULATION ---
                sr_no = 1
                sum_taxable = sum_cgst = sum_sgst = sum_igst = sum_cess = sum_total_itc = 0.0

                for index, row in df_gstr2b.iterrows():
                    curr = start_row + index
                    
                    taxable = get_amount(row[col_taxable] if col_taxable else 0)
                    cgst = get_amount(row[col_cgst] if col_cgst else 0)
                    sgst = get_amount(row[col_sgst] if col_sgst else 0)
                    igst = get_amount(row[col_igst] if col_igst else 0)
                    cess = get_amount(row[col_cess] if col_cess else 0)
                    row_itc = cgst + sgst + igst + cess
                    
                    sum_taxable += taxable
                    sum_cgst += cgst
                    sum_sgst += sgst
                    sum_igst += igst
                    sum_cess += cess
                    sum_total_itc += row_itc

                    data_map = {
                        1: sr_no, 2: row[col_gstin] if col_gstin else '', 
                        3: row[col_name] if col_name else '', 4: row[col_inv] if col_inv else '', 
                        5: row[col_date] if col_date else '',
                        10: taxable, 11: "Input", 13: cgst, 14: sgst, 
                        15: igst, 16: cess, 17: "Yes", 18: row_itc, 19: "Y"
                    }

                    for col_idx in range(1, 20):
                        cell = sheet.cell(row=curr, column=col_idx)
                        cell.value = data_map.get(col_idx, '')
                        cell.border = thin_border
                        cell.font = normal_font 
                        
                        if col_idx == 3: 
                            cell.alignment = align_left
                        elif col_idx in [10, 13, 14, 15, 16, 18]: 
                            cell.alignment = align_right
                            if cell.value != '': cell.number_format = '#,##0.00' 
                        else:
                            cell.alignment = align_center

                    sr_no += 1

                # --- 5. ULTRA CLEANUP & FIXING TOTALS ---
                data_end_row = start_row + num_rows - 1
                search_start = data_end_row + 1

                # 5A. Purana TOTAL dhoondhna aur kachra saaf karna
                total_row_idx = None
                for r in range(search_start, sheet.max_row + 100):
                    val = str(sheet.cell(row=r, column=1).value).strip().upper()
                    if "TOTAL" in val:
                        total_row_idx = r
                        break

                if total_row_idx:
                    gap = total_row_idx - search_start
                    if gap > 0:
                        sheet.delete_rows(search_start, amount=gap)

                # 5B. Naya Total likhna
                amounts_map = {10: sum_taxable, 13: sum_cgst, 14: sum_sgst, 15: sum_igst, 16: sum_cess, 18: sum_total_itc}
                for col_idx, val in amounts_map.items():
                    cell = sheet.cell(row=search_start, column=col_idx)
                    cell.value = val 
                    cell.font = bold_font 
                    cell.alignment = align_right
                    cell.number_format = '#,##0.00'

                # --- 6. SMART SIGNATURE BLOCK FINDER & HARD TRIM ---
                
                # Sheet me "PARTNER" ya "KALUBHAI" kis line par likha hai, usko dhoondhna
                partner_row_idx = search_start 
                
                for r in range(search_start, sheet.max_row + 50):
                    for c in range(1, 10): # Column A se I tak check karega
                        val = str(sheet.cell(row=r, column=c).value).strip().upper()
                        if "PARTNER" in val or "KALUBHAI" in val:
                            if r > partner_row_idx:
                                partner_row_idx = r
                
                # Target Last Row = Jaha "PARTNER" mila, theek uske 2 khali line baad
                target_last_row = partner_row_idx + 2 

                # 🚀 SMART HARD TRIM: Delete EVERYTHING below target_last_row
                if sheet.max_row > target_last_row:
                    sheet.delete_rows(target_last_row + 1, amount=2000) # Niche ki 2000 lines uda dega taaki extra pages na aayein

                # Print Area ko exactly wahi tak Lock karna
                sheet.print_area = f"A1:S{target_last_row}"

                # --- 7. File Download Ready ---
                output = io.BytesIO()
                wb.save(output)
                output.seek(0) 

                st.success("✅ Success! Ab 'PARTNER' ka naam ekdum safe hai aur extra pages completely remove ho gaye hain.")
                
                st.download_button(
                    label="📥 Download Perfectly Trimmed Annexure-B",
                    data=output,
                    file_name="Annexure_B_Smart_Trim.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary"
                )

            except Exception as e:
                st.error(f"❌ Kuch error aayi: {e}")
