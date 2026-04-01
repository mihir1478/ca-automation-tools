import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
import io

# --- Page Settings ---
st.set_page_config(page_title="Professional GSTR-2B Filler", page_icon="🏢", layout="wide")

# --- 🧠 MASTER LOGIC: Firms & Mappings ---
all_firm_logics = {
    "BHAVANI EXPORTS": {
        "partner_name": "KALUBHAI RAMJIBHAI JASOLIYA",
        "mappings": {
            "FINESTAR JEWELLERY & DIAMONDS PRIVATE LIMITED": {"cat": "Input Goods", "hsn": "71023910"},
            "ROCK CANDY IMPEX": {"cat": "Input Goods", "hsn": "71023910"},
            "NAROLA DIAMONDS PRIVATE LIMITED": {"cat": "Input Goods", "hsn": "71023910"},
            "KAYRA CRYSTAL LLP": {"cat": "Input Goods", "hsn": "71023910"},
            "ICICI BANK LIMITED": {"cat": "Input Services", "hsn": "997111"},
        }
    }
}

st.sidebar.title("🏢 Firm Selection")
selected_firm = st.sidebar.selectbox("Select Firm", list(all_firm_logics.keys()))
firm_info = all_firm_logics[selected_firm]
firm_mapping = firm_info["mappings"]

st.title(f"📊 Annexure-B Auto-Filler: {selected_firm}")
st.write("Ab Signature Area ekdum saaf aayega, bina kisi border ke!")

template_file = st.file_uploader(f"1. Upload Annexure-B Template (.xlsx)", type=['xlsx'])
gstr2b_file = st.file_uploader("2. Upload GSTR-2B File (.xlsx)", type=['xlsx', 'xls'])

if st.button("🚀 Process & Create Perfect File", type="primary", use_container_width=True):
    if template_file and gstr2b_file:
        with st.spinner("⏳ Formatting Signature Area..."):
            try:
                # --- 1. Load GSTR-2B ---
                df_raw = pd.read_excel(gstr2b_file, sheet_name='B2B', header=None)
                header_idx = 5 
                for i, r in df_raw.head(15).iterrows():
                    vals = str(r.values).lower()
                    if 'gstin' in vals and ('trade' in vals or 'name' in vals):
                        header_idx = i
                        break
                df_gstr2b = pd.read_excel(gstr2b_file, sheet_name='B2B', header=header_idx).fillna('')
                cols = [str(c).lower().replace(' ', '').replace('\n', '').replace('/', '') for c in df_gstr2b.columns]
                df_gstr2b.columns = cols

                # Mapping Column Names
                def f_c(keys, d_idx):
                    for k in keys:
                        for i, c in enumerate(cols):
                            if k in c: return c
                    return cols[d_idx] if len(cols) > d_idx else None

                c_gst, c_nm, c_inv, c_dt, c_tx = f_c(['gstin'],0), f_c(['trade','legal','name'],1), f_c(['inv'],2), f_c(['date'],4), f_c(['taxable'],9)
                c_ig, c_cg, c_sg, c_ce = f_c(['integrated','igst'],10), f_c(['central','cgst'],11), f_c(['state','sgst'],12), f_c(['cess'],13)

                def get_amt(val):
                    try: return float(str(val).replace(',', '').strip()) if str(val).strip() != '' else 0.0
                    except: return 0.0

                # --- 2. Excel Setup ---
                wb = openpyxl.load_workbook(template_file)
                sheet = wb.active
                
                # Global Styles
                thin = Side(style='thin')
                border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
                no_border = Border() # 👈 Sabse important: Khali border
                f_cal = Font(name='Calibri', size=11)
                f_bold = Font(name='Calibri', size=11, bold=True)
                
                # PDF Settings
                sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
                sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
                sheet.sheet_properties.pageSetUpPr.fitToPage = True
                sheet.page_setup.fitToWidth = 1
                sheet.page_setup.fitToHeight = 0 
                sheet.page_margins = PageMargins(left=0.2, right=0.2, top=0.4, bottom=0.4)

                # --- 3. Insert Data & Keep Template Safe ---
                start_row = 6
                num_rows = len(df_gstr2b)
                sheet.insert_rows(start_row, amount=num_rows)

                s_tx = s_cg = s_sg = s_ig = s_ce = s_itc = 0.0
                for index, row in df_gstr2b.iterrows():
                    curr = start_row + index
                    name = str(row[c_nm]).strip()
                    m_key = name.upper()
                    cat = firm_mapping.get(m_key, {}).get("cat", "Input Goods")
                    hsn = firm_mapping.get(m_key, {}).get("hsn", "71023910")

                    tx = get_amt(row[c_tx]); cg = get_amt(row[c_cg]); sg = get_amt(row[c_sg])
                    ig = get_amt(row[c_ig]); ce = get_amt(row[c_ce]); ritc = cg+sg+ig+ce
                    s_tx+=tx; s_cg+=cg; s_sg+=sg; s_ig+=ig; s_ce+=ce; s_itc+=ritc

                    v_map = {1:index+1, 2:row[c_gst], 3:name, 4:row[c_inv], 5:row[c_dt], 
                             10:tx, 11:cat, 12:hsn, 13:cg, 14:sg, 15:ig, 16:ce, 17:"Yes", 18:ritc, 19:"Y"}

                    for c_idx in range(1, 20):
                        cell = sheet.cell(row=curr, column=c_idx)
                        cell.value = v_map.get(c_idx, '')
                        cell.border = border_all
                        cell.font = f_cal
                        if c_idx == 3: cell.alignment = Alignment(horizontal='left')
                        elif c_idx in [10, 13, 14, 15, 16, 18]: 
                            cell.alignment = Alignment(horizontal='right')
                            cell.number_format = '#,##0.00'
                        else: cell.alignment = Alignment(horizontal='center')

                # --- 4. TOTAL Row Fix ---
                new_start = start_row + num_rows
                total_row_idx = None
                for r in range(new_start, sheet.max_row + 1):
                    if "TOTAL" in str(sheet.cell(row=r, column=1).value).upper():
                        total_row_idx = r
                        break

                if total_row_idx:
                    # Gap delete karna (Extra space removal)
                    gap = total_row_idx - new_start
                    if gap > 0: sheet.delete_rows(new_start, amount=gap)
                    
                    # TOTAL Row Pe Borders Lagana
                    t_map = {10: s_tx, 13: s_cg, 14: s_sg, 15: s_ig, 16: s_ce, 18: s_itc}
                    for c_idx in range(1, 20):
                        c = sheet.cell(row=new_start, column=c_idx)
                        c.border = border_all # TOTAL tak borders rahenge
                        if c_idx in t_map:
                            c.value = t_map[c_idx]
                            c.font = f_bold; c.number_format = '#,##0.00'; c.alignment = Alignment(horizontal='right')
                        elif c_idx == 1:
                            c.value = "TOTAL"; c.font = f_bold; c.alignment = Alignment(horizontal='center')

                # --- 5. CLEAN SIGNATURE AREA (image_3c4e87.png Style) ---
                # TOTAL ke theek niche se borders mita dena
                sig_start_row = new_start + 1
                for r in range(sig_start_row, sig_start_row + 20):
                    for c_idx in range(1, 20):
                        sheet.cell(row=r, column=c_idx).border = no_border # Borders Hamesha ke liye Gayab

                # Ab Signature ke details ko "Clean" cells me daalna
                sig_base = new_start + 2
                # Line 1: FOR, FIRM NAME
                sheet.cell(row=sig_base, column=2).value = f"FOR, {selected_firm}"
                sheet.cell(row=sig_base, column=2).font = f_bold
                
                # Line 2: Name (Signature Gap ke baad)
                sheet.cell(row=sig_base + 4, column=2).value = firm_info["partner_name"]
                sheet.cell(row=sig_base + 4, column=2).font = f_bold
                
                # Line 3: Designation
                sheet.cell(row=sig_base + 5, column=2).value = "PARTNER"
                sheet.cell(row=sig_base + 5, column=2).font = f_bold

                # Print Area Set (No Extra Pages)
                final_row = sig_base + 7
                sheet.print_area = f"A1:S{final_row}"

                # --- 6. Export ---
                out = io.BytesIO()
                wb.save(out)
                out.seek(0)
                st.success("✅ Signature block ab ekdum clean (Bina Border ke) hai!")
                st.download_button(f"📥 Download Perfect File", out, f"{selected_firm}_Annexure_B.xlsx")

            except Exception as e:
                st.error(f"❌ Error: {e}")
